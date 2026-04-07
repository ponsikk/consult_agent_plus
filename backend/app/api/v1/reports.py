from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.deps import get_current_user
from app.models.analysis import Analysis, AnalysisPhoto, Defect, DefectType
from app.models.user import User
from app.schemas.analysis import DefectTypeOut
from app.services.pdf_service import generate_pdf
from app.services.storage_service import storage_service

router = APIRouter(tags=["reports"])


@router.get("/analyses/{analysis_id}/report/pdf")
async def download_pdf_report(
    analysis_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(Analysis)
        .where(Analysis.id == analysis_id, Analysis.user_id == current_user.id)
        .options(
            selectinload(Analysis.photos).selectinload(AnalysisPhoto.defects).selectinload(Defect.defect_type)
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")
    if analysis.status != "done":
        raise HTTPException(status_code=400, detail="Report not ready: analysis still processing")

    photos_data = []
    for idx, photo in enumerate(analysis.photos):
        defects_data = []
        for d in photo.defects:
            defects_data.append({
                "criticality": d.criticality,
                "type_code": d.defect_type.code if d.defect_type else "—",
                "type_name": d.defect_type.name if d.defect_type else "Неизвестный тип",
                "description": d.description,
                "consequences": d.consequences,
                "recommendations": d.recommendations,
                "norm_references": d.norm_references or [],
            })

        # Скачиваем аннотированное фото (или оригинал если аннотация не сформирована)
        image_bytes: bytes | None = None
        try:
            key = photo.annotated_key or photo.original_key
            image_bytes = await storage_service.download_file(key)
        except Exception:
            pass

        photos_data.append({"index": idx, "defects": defects_data, "image_bytes": image_bytes})

    data = {
        "analysis": {
            "object_name": analysis.object_name,
            "shot_date": analysis.shot_date.strftime("%d.%m.%Y") if analysis.shot_date else "—",
            "created_at": analysis.created_at.strftime("%d.%m.%Y %H:%M") if analysis.created_at else "—",
            "completed_at": analysis.completed_at.strftime("%d.%m.%Y %H:%M") if analysis.completed_at else None,
        },
        "user": {
            "full_name": current_user.full_name,
            "email": current_user.email,
        },
        "photos": photos_data,
    }

    pdf_bytes = await generate_pdf(data)

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="report_{analysis_id}.pdf"'},
    )


@router.get("/analyses/{analysis_id}/report/excel")
async def download_excel_report(
    analysis_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(Analysis)
        .where(Analysis.id == analysis_id, Analysis.user_id == current_user.id)
        .options(
            selectinload(Analysis.photos).selectinload(AnalysisPhoto.defects).selectinload(Defect.defect_type)
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")
    if analysis.status != "done":
        raise HTTPException(status_code=400, detail="Report not ready: analysis still processing")

    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Дефекты"

    headers = ["№", "Фото", "Код типа", "Тип дефекта", "Критичность", "Описание", "Последствия", "Рекомендации", "Нормативы"]
    col_widths = [5, 6, 12, 25, 14, 40, 35, 40, 30]
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[cell.column_letter].width = w

    CRITICALITY_RU = {"critical": "Критический", "significant": "Значительный", "minor": "Незначительный"}
    row = 2
    for idx, photo in enumerate(analysis.photos):
        for d in photo.defects:
            ws.append([
                row - 1,
                idx + 1,
                d.defect_type.code if d.defect_type else "—",
                d.defect_type.name if d.defect_type else "—",
                CRITICALITY_RU.get(d.criticality, d.criticality),
                d.description,
                d.consequences,
                d.recommendations,
                ", ".join(d.norm_references or []) or "—",
            ])
            for col_idx in range(1, 10):
                ws.cell(row=row, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="report_{analysis_id}.xlsx"'},
    )


@router.get("/defects/catalog", response_model=List[DefectTypeOut])
async def get_defect_catalog(
    system: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(DefectType)
    if system:
        query = query.where(DefectType.system == system)
    result = await db.execute(query.order_by(DefectType.system, DefectType.code))
    return result.scalars().all()
